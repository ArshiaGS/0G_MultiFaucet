from typing import Optional
from hdwallet import BIP44HDWallet
from hdwallet.cryptocurrencies import EthereumMainnet
from hdwallet.derivations import BIP44Derivation
from hdwallet.utils import generate_mnemonic
import openpyxl

book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = "Address"
sheet['B1'] = "Private key"
sheet['C1'] = "Mnemonic"

# Number of wallets
N = int(input("Enter number of wallets: "))

for i in range(N):
    # Generate english mnemonic words
    MNEMONIC: str = generate_mnemonic(language="english", strength=128)
    # Secret passphrase/password for mnemonic
    PASSPHRASE: Optional[str] = None  # "meherett"

    # Initialize Ethereum mainnet BIP44HDWallet
    bip44_hdwallet: BIP44HDWallet = BIP44HDWallet(cryptocurrency=EthereumMainnet)
    # Get Ethereum BIP44HDWallet from mnemonic
    bip44_hdwallet.from_mnemonic(
        mnemonic=MNEMONIC, language="english", passphrase=PASSPHRASE
    )
    # Clean default BIP44 derivation indexes/paths
    bip44_hdwallet.clean_derivation()
    sheet.cell(row=i+2,column=3).value=bip44_hdwallet.mnemonic()


    # Get Ethereum BIP44HDWallet information's from address index

    # Derivation from Ethereum BIP44 derivation path
    bip44_derivation: BIP44Derivation = BIP44Derivation(
        cryptocurrency=EthereumMainnet, account=0, change=False, address=0
    )
    # Drive Ethereum BIP44HDWallet
    bip44_hdwallet.from_path(path=bip44_derivation)
    # Print address_index, path, address and private_key
    sheet.cell(row=i+2, column=1).value=bip44_hdwallet.address()
    sheet.cell(row=i+2, column=2).value=bip44_hdwallet.private_key()
    # Clean derivation indexes/paths
    bip44_hdwallet.clean_derivation()

book.save("new_wallets.xlsx")
book.close()
